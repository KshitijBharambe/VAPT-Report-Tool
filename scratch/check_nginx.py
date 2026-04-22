"""Check why nginx version detection not triggering."""

import openpyxl, re

xlsx = "tests/input/AEG_Subnet1_y06m7b 1.xlsx"
wb = openpyxl.load_workbook(xlsx)
ws = wb[wb.sheetnames[0]]
headers = [cell.value for cell in ws[1]]
col = {h: i for i, h in enumerate(headers) if h}

for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    name = str(row[col["Name"]] or "")
    if "nginx" in name.lower():
        host = row[col["Host"]]
        port = row[col["Port"]]
        risk = row[col["Risk"]]
        plugin_output = str(row[col["Plugin Output"]] or "")

        print(f"Name: {name}")
        print(f"Host: {host}:{port}")
        print(f"Risk: {risk}")
        print(f"Plugin Output (first 500 chars): {plugin_output[:500]}")

        # Try the version regex
        version_match = re.search(r"version\s*:\s*(\d[\d.]+)", plugin_output, re.I)
        if version_match:
            print(f"Version found: {version_match.group(1)}")
            from generate_report import _is_version_less_than

            print(
                f"Is < 1.17.7: {_is_version_less_than(version_match.group(1), '1.17.7')}"
            )
        else:
            print("No version match in plugin output")
        print("---")
