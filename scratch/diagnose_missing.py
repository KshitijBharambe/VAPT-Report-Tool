"""Diagnose why specific vulns are missing from generated report."""

import openpyxl, re

xlsx = "tests/input/AEG_Subnet1_y06m7b 1.xlsx"
wb = openpyxl.load_workbook(xlsx)
ws = wb[wb.sheetnames[0]]
headers = [cell.value for cell in ws[1]]
col = {h: i for i, h in enumerate(headers) if h}

# 1. Check for SNMP
print("=== SNMP vulns in xlsx ===")
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    name = str(row[col["Name"]] or "")
    if "snmp" in name.lower():
        print(
            f"  [{row[col['Risk']]}] {name} — Host: {row[col['Host']]}:{row[col['Port']]}"
        )

# 2. Check for NCache
print("\n=== NCache / ncache vulns in xlsx ===")
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    name = str(row[col["Name"]] or "")
    if "ncache" in name.lower() or "cache" in name.lower():
        print(
            f"  [{row[col['Risk']]}] {name} — Host: {row[col['Host']]}:{row[col['Port']]}"
        )

# 3. Check for nginx
print("\n=== nginx vulns in xlsx ===")
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    name = str(row[col["Name"]] or "")
    if "nginx" in name.lower():
        host = row[col["Host"]]
        port = row[col["Port"]]
        risk = row[col["Risk"]]
        plugin_output = str(row[col.get("Plugin Output", -1)] or "")[:200]
        print(f"  [{risk}] {name} — Host: {host}:{port}")
        if "version" in plugin_output.lower():
            print(f"    Plugin Output: {plugin_output}")

# 4. Check what "Informational/None" vulns exist that SHOULD be included
print("\n=== Risk=None vulns that handmade report includes ===")
interesting_none = []
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    name = str(row[col["Name"]] or "")
    risk = str(row[col["Risk"]] or "")
    if risk in ("None", "") and any(
        kw in name.lower() for kw in ("snmp", "ncache", "cache", "exposed", "service")
    ):
        interesting_none.append((risk, name, row[col["Host"]]))
for r, n, h in interesting_none[:20]:
    print(f"  [{r}] {n} — {h}")

# 5. All unique Risk values
from collections import Counter

risks = Counter()
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    risks[row[col["Risk"]]] += 1
print(f"\n=== Risk distribution ===\n{dict(risks)}")

# 6. False Positive Check column
fp_col = col.get("False Positive Check")
if fp_col is not None:
    fp_vals = Counter()
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        fp_vals[row[fp_col]] += 1
    print(f"\n=== False Positive Check distribution ===\n{dict(fp_vals)}")
